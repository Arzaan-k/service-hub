import openpyxl
from datetime import datetime
from collections import Counter

file_path = r'c:\Users\msi user\Desktop\Developer\service-hub-ui\service-hub-ui\Serivce Master.xlsx'

wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb['Sheet1']

print('='*100)
print('DATA PATTERN ANALYSIS - Serivce Master.xlsx')
print('='*100)
print()

# Get headers
headers = []
for col in range(1, ws.max_column + 1):
    cell_value = ws.cell(1, col).value
    headers.append(cell_value if cell_value else f'Column_{col}')

# Analyze unique values in key categorical fields
categorical_fields = {
    13: 'Machine Status',
    22: 'Machine Make',
    23: 'Work Type',
    24: 'Client Type',
    25: 'Job Type',
    73: 'Service Type',
    79: 'Call Attended Type',
    21: 'Container Size',
    32: 'Reefer Unit',
    37: 'Brand new / Used',
    52: 'Billing Type',
    65: 'Required Material Sent Through',
    44: 'Indent Required ?',
    56: 'Spares Required ?',
}

print('CATEGORICAL FIELD VALUE ANALYSIS')
print('='*100)

for col_idx, field_name in categorical_fields.items():
    values = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, col_idx).value
        if val:
            values.append(str(val).strip())

    if values:
        value_counts = Counter(values)
        print(f'\n{field_name} (Column {col_idx}):')
        print(f'  Total unique values: {len(value_counts)}')
        print(f'  Value distribution:')
        for value, count in value_counts.most_common(10):
            percentage = (count / len(values)) * 100
            print(f'    - {value}: {count} ({percentage:.1f}%)')

# Analyze technician distribution
print('\n\n' + '='*100)
print('TECHNICIAN ANALYSIS')
print('='*100)

techs = []
for row in range(2, ws.max_row + 1):
    val = ws.cell(row, 39).value
    if val:
        techs.append(str(val).strip().upper())

tech_counts = Counter(techs)
print(f'\nTotal unique technicians: {len(tech_counts)}')
print(f'Top 15 technicians by job count:')
for tech, count in tech_counts.most_common(15):
    print(f'  - {tech}: {count} jobs')

# Analyze client distribution
print('\n\n' + '='*100)
print('CLIENT ANALYSIS')
print('='*100)

clients = []
for row in range(2, ws.max_row + 1):
    val = ws.cell(row, 4).value
    if val:
        clients.append(str(val).strip())

client_counts = Counter(clients)
print(f'\nTotal unique clients: {len(client_counts)}')
print(f'Top 20 clients by service count:')
for client, count in client_counts.most_common(20):
    print(f'  - {client}: {count} services')

# Analyze container numbers
print('\n\n' + '='*100)
print('CONTAINER NUMBER ANALYSIS')
print('='*100)

containers = []
for row in range(2, ws.max_row + 1):
    val = ws.cell(row, 8).value
    if val:
        containers.append(str(val).strip().upper())

container_counts = Counter(containers)
print(f'\nTotal unique containers: {len(container_counts)}')
print(f'Containers with multiple service records (Top 20):')
multi_service = [(cont, count) for cont, count in container_counts.most_common() if count > 1]
for cont, count in multi_service[:20]:
    print(f'  - {cont}: {count} services')

# Analyze container prefixes
print('\n\nContainer Owner Code Distribution:')
prefixes = [cont[:4] for cont in containers if len(cont) >= 4]
prefix_counts = Counter(prefixes)
for prefix, count in prefix_counts.most_common(15):
    percentage = (count / len(prefixes)) * 100
    print(f'  - {prefix}: {count} ({percentage:.1f}%)')

# Analyze date ranges
print('\n\n' + '='*100)
print('DATE RANGE ANALYSIS')
print('='*100)

service_dates = []
for row in range(2, ws.max_row + 1):
    val = ws.cell(row, 72).value  # Complaint Attended Date
    if val and isinstance(val, datetime):
        service_dates.append(val)

if service_dates:
    print(f'\nService Date Range:')
    print(f'  Earliest service: {min(service_dates).strftime("%Y-%m-%d")}')
    print(f'  Latest service: {max(service_dates).strftime("%Y-%m-%d")}')
    print(f'  Total services with dates: {len(service_dates)}')

    # Monthly distribution
    monthly = Counter([d.strftime('%Y-%m') for d in service_dates])
    print(f'\n  Monthly service distribution:')
    for month, count in sorted(monthly.items()):
        print(f'    - {month}: {count} services')

# Analyze complaint types
print('\n\n' + '='*100)
print('COMPLAINT TYPE ANALYSIS')
print('='*100)

complaints = []
for row in range(2, ws.max_row + 1):
    val = ws.cell(row, 9).value  # What's the complaint
    if val:
        complaints.append(str(val).strip())

complaint_counts = Counter(complaints)
print(f'\nTotal unique complaint types: {len(complaint_counts)}')
print(f'Top 20 most common complaints:')
for complaint, count in complaint_counts.most_common(20):
    percentage = (count / len(complaints)) * 100
    print(f'  - {complaint[:60]}: {count} ({percentage:.1f}%)')

# Analyze issues found
print('\n\n' + '='*100)
print('ISSUES FOUND ANALYSIS')
print('='*100)

issues = []
for row in range(2, ws.max_row + 1):
    val = ws.cell(row, 26).value  # Issue(s) found
    if val:
        issues.append(str(val).strip())

issue_counts = Counter(issues)
print(f'\nTotal unique issues: {len(issue_counts)}')
print(f'Top 20 most common issues:')
for issue, count in issue_counts.most_common(20):
    percentage = (count / len(issues)) * 100
    issue_text = issue[:80] if len(issue) > 80 else issue
    print(f'  - {issue_text}: {count} ({percentage:.1f}%)')

# Analyze equipment inspection results
print('\n\n' + '='*100)
print('EQUIPMENT CONDITION ANALYSIS')
print('='*100)

inspection_fields = {
    84: 'Condenser Coil',
    87: 'Evaporator Coil',
    89: 'Compressor Oil',
    91: 'Refrigerant Gas',
    93: 'Controller Display',
    95: 'Controller keypad',
    100: 'Compressor contactor',
    102: 'EVP/COND contactor',
}

for col_idx, field_name in inspection_fields.items():
    values = []
    for row in range(2, min(500, ws.max_row + 1)):  # Sample first 500 rows
        val = ws.cell(row, col_idx).value
        if val:
            values.append(str(val).strip())

    if values:
        value_counts = Counter(values)
        print(f'\n{field_name}:')
        for value, count in value_counts.most_common(5):
            percentage = (count / len(values)) * 100
            print(f'  - {value}: {count} ({percentage:.1f}%)')

# Location analysis
print('\n\n' + '='*100)
print('LOCATION ANALYSIS')
print('='*100)

locations = []
for row in range(2, ws.max_row + 1):
    val = ws.cell(row, 76).value  # Client Location
    if val:
        locations.append(str(val).strip())

location_counts = Counter(locations)
print(f'\nTotal unique locations: {len(location_counts)}')
print(f'Top 20 service locations:')
for location, count in location_counts.most_common(20):
    percentage = (count / len(locations)) * 100
    print(f'  - {location}: {count} services ({percentage:.1f}%)')

print('\n\n' + '='*100)
print('ANALYSIS COMPLETE')
print('='*100)
