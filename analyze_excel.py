import openpyxl
from datetime import datetime

file_path = r'c:\Users\msi user\Desktop\Developer\service-hub-ui\service-hub-ui\Serivce Master.xlsx'

wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb['Sheet1']

# Get all headers
headers = []
for col in range(1, ws.max_column + 1):
    cell_value = ws.cell(1, col).value
    headers.append(cell_value if cell_value else f'Column_{col}')

print('='*100)
print('COMPREHENSIVE DATA STRUCTURE ANALYSIS - Serivce Master.xlsx')
print('='*100)
print()
print(f'Total Records: {ws.max_row - 1}')
print(f'Total Columns: {ws.max_column}')
print()

# Analyze data types and patterns
print('='*100)
print('DETAILED COLUMN ANALYSIS')
print('='*100)

for idx, header in enumerate(headers, 1):
    col_idx = idx

    # Sample non-empty values
    sample_values = []
    data_types = set()
    non_empty_count = 0

    for row in range(2, min(20, ws.max_row + 1)):
        value = ws.cell(row, col_idx).value
        if value is not None and str(value).strip():
            non_empty_count += 1
            if len(sample_values) < 5:
                sample_values.append(value)
            data_types.add(type(value).__name__)

    print(f'\n{idx}. {header}')
    dt_str = ', '.join(data_types) if data_types else 'No data'
    print(f'   Data Types: {dt_str}')
    print(f'   Sample Values:')
    for sv in sample_values[:3]:
        if isinstance(sv, datetime):
            print(f'      - {sv.strftime("%Y-%m-%d %H:%M:%S")}')
        else:
            val_str = str(sv)[:100]
            print(f'      - {val_str}')

print('\n' + '='*100)
print('KEY FIELD IDENTIFICATION')
print('='*100)

# Identify Job Order column
print('\nJOB ORDER FIELD:')
print(f'  Column 1: {headers[0]}')
job_orders = []
for row in range(2, min(10, ws.max_row + 1)):
    val = ws.cell(row, 1).value
    if val:
        job_orders.append(val)
print(f'  Sample Job Orders: {", ".join(str(j) for j in job_orders[:5])}')

# Identify Container Number columns
print('\nCONTAINER NUMBER FIELDS:')
container_cols = []
for idx, header in enumerate(headers, 1):
    h_lower = str(header).lower()
    if 'container' in h_lower and ('number' in h_lower or 'no' in h_lower):
        container_cols.append((idx, header))

for col_idx, col_name in container_cols:
    samples = []
    for row in range(2, min(10, ws.max_row + 1)):
        val = ws.cell(row, col_idx).value
        if val:
            samples.append(val)
    print(f'  Column {col_idx}: {col_name}')
    sample_str = ', '.join(str(s) for s in samples[:5])
    print(f'    Samples: {sample_str}')

# Identify Date fields
print('\nDATE FIELDS:')
date_cols = []
for idx, header in enumerate(headers, 1):
    h_lower = str(header).lower()
    if 'date' in h_lower or 'timestamp' in h_lower:
        date_cols.append((idx, header))

for col_idx, col_name in date_cols[:10]:
    samples = []
    for row in range(2, min(6, ws.max_row + 1)):
        val = ws.cell(row, col_idx).value
        if val:
            if isinstance(val, datetime):
                samples.append(val.strftime('%Y-%m-%d'))
            else:
                samples.append(str(val))
    print(f'  Column {col_idx}: {col_name}')
    if samples:
        sample_str = ', '.join(samples[:3])
        print(f'    Samples: {sample_str}')

# Identify Cost/Price fields
print('\nCOST/PRICE/BILLING FIELDS:')
cost_cols = []
for idx, header in enumerate(headers, 1):
    h = str(header).lower()
    if any(word in h for word in ['cost', 'price', 'amount', 'billing', 'po', 'payment', 'indent']):
        cost_cols.append((idx, header))

for col_idx, col_name in cost_cols:
    print(f'  Column {col_idx}: {col_name}')

# Identify Technician fields
print('\nTECHNICIAN FIELDS:')
tech_cols = []
for idx, header in enumerate(headers, 1):
    if 'technician' in str(header).lower():
        tech_cols.append((idx, header))

for col_idx, col_name in tech_cols:
    samples = []
    for row in range(2, min(8, ws.max_row + 1)):
        val = ws.cell(row, col_idx).value
        if val and val not in samples:
            samples.append(val)
    print(f'  Column {col_idx}: {col_name}')
    if samples:
        sample_str = ', '.join(str(s) for s in samples[:5])
        print(f'    Samples: {sample_str}')

# Identify Client fields
print('\nCLIENT FIELDS:')
client_cols = []
for idx, header in enumerate(headers, 1):
    if 'client' in str(header).lower():
        client_cols.append((idx, header))

for col_idx, col_name in client_cols:
    samples = []
    for row in range(2, min(6, ws.max_row + 1)):
        val = ws.cell(row, col_idx).value
        if val:
            samples.append(val)
    print(f'  Column {col_idx}: {col_name}')
    if samples:
        sample_str = ', '.join(str(s)[:30] for s in samples[:3])
        print(f'    Samples: {sample_str}')

# Identify Equipment/Machine fields
print('\nEQUIPMENT/MACHINE FIELDS:')
equip_cols = []
for idx, header in enumerate(headers, 1):
    h = str(header).lower()
    if any(word in h for word in ['reefer', 'machine', 'compressor', 'controller', 'evaporator', 'condenser', 'model', 'serial']):
        equip_cols.append((idx, header))

for col_idx, col_name in equip_cols[:15]:
    print(f'  Column {col_idx}: {col_name}')

# Identify Issue/Complaint fields
print('\nISSUE/COMPLAINT FIELDS:')
issue_cols = []
for idx, header in enumerate(headers, 1):
    h = str(header).lower()
    if any(word in h for word in ['complaint', 'issue', 'problem', 'alarm', 'description']):
        issue_cols.append((idx, header))

for col_idx, col_name in issue_cols:
    print(f'  Column {col_idx}: {col_name}')

# Identify Spare Parts fields
print('\nSPARE PARTS FIELDS:')
spare_cols = []
for idx, header in enumerate(headers, 1):
    h = str(header).lower()
    if any(word in h for word in ['spare', 'part', 'material', 'required']):
        spare_cols.append((idx, header))

for col_idx, col_name in spare_cols[:15]:
    print(f'  Column {col_idx}: {col_name}')

# Identify Inspection fields
print('\nINSPECTION/CONDITION FIELDS:')
inspect_cols = []
for idx, header in enumerate(headers, 1):
    h = str(header).lower()
    if any(word in h for word in ['condition', 'coil', 'motor', 'oil', 'gas', 'display', 'keypad', 'cable', 'contactor', 'mcb', 'filter', 'pressure', 'current', 'voltage', 'pti']):
        inspect_cols.append((idx, header))

for col_idx, col_name in inspect_cols[:20]:
    print(f'  Column {col_idx}: {col_name}')

print()
print('='*100)
print('SUMMARY')
print('='*100)
print(f'Total columns analyzed: {len(headers)}')
print(f'Container number fields: {len(container_cols)}')
print(f'Date/Timestamp fields: {len(date_cols)}')
print(f'Client-related fields: {len(client_cols)}')
print(f'Technician fields: {len(tech_cols)}')
print(f'Equipment/Machine fields: {len(equip_cols)}')
print(f'Issue/Complaint fields: {len(issue_cols)}')
print(f'Spare parts fields: {len(spare_cols)}')
print(f'Inspection fields: {len(inspect_cols)}')
print()
